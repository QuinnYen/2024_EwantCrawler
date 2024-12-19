import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from PyQt6.QtWidgets import QTableWidget, QFileDialog, QMessageBox

class CourseExporter:
    def __init__(self, table_widget: QTableWidget):
        self.table_widget = table_widget
    
    def export_to_excel(self) -> bool:
        """
        將課程資料匯出到 Excel
        Returns:
            bool: 是否匯出成功
        """
        try:
            # 檢查是否有資料可匯出
            if self.table_widget.rowCount() == 0:
                QMessageBox.warning(None, "警告", "沒有資料可供匯出！")
                return False
            
            # 選擇儲存路徑
            file_path, _ = QFileDialog.getSaveFileName(
                None, 
                "匯出 Excel 檔案", 
                "", 
                "Excel 檔案 (*.xlsx)"
            )
            if not file_path:
                return False
                
            # 確保檔案副檔名為 .xlsx
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            
            # 建立新的 Excel 活頁簿和工作表
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "課程資料"
            
            # 設定標題列
            headers = [
                "課程狀態",
                "課程名稱",
                "開始時間",
                "結束時間",
                "選修人數(台灣)",
                "選修人數(中國大陸)",
                "選修人數(其他)",
                "通過人數(台灣)",
                "通過人數(中國大陸)",
                "通過人數(其他)",
                "影片瀏覽次數(台灣)",
                "影片瀏覽次數(中國大陸)",
                "影片瀏覽次數(其他)",
                "作業測驗作答次數(台灣)",
                "作業測驗作答次數(中國大陸)",
                "作業測驗作答次數(其他)",
                "講義參考資料瀏覽次數(台灣)",
                "講義參考資料瀏覽次數(中國大陸)",
                "講義參考資料瀏覽次數(其他)",
                "討論次數"
            ]
            sheet.append(headers)
            
            # 設定標題列樣式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for idx, cell in enumerate(sheet[1]):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                # 設定欄寬
                if idx == 1:  # 課程名稱欄位
                    sheet.column_dimensions[cell.column_letter].width = 40
                else:  # 其他欄位
                    sheet.column_dimensions[cell.column_letter].width = 15
            
            # 寫入課程資料
            row_count = self.table_widget.rowCount()
            for row in range(row_count):
                course_data = []
                for col in range(self.table_widget.columnCount()):
                    item = self.table_widget.item(row, col)
                    if col >= 4:  # 數字欄位
                        try:
                            original_value = item.text() if item else ""
                            # 嘗試轉換為數字，如果失敗就保留原始值
                            try:
                                value = int(original_value)
                                course_data.append(value)
                            except ValueError:
                                course_data.append(original_value)
                        except:
                            course_data.append("")
                    else:  # 文字欄位
                        course_data.append(item.text() if item else "")
                sheet.append(course_data)
            
            # 設定資料列樣式
            data_alignment_right = Alignment(horizontal='right', vertical='center')
            data_alignment_left = Alignment(horizontal='left', vertical='center')
            
            for row in sheet.iter_rows(min_row=2):
                for idx, cell in enumerate(row):
                    if idx == 1:  # 課程名稱靠左
                        cell.alignment = data_alignment_left
                    else:  # 其他欄位靠右
                        cell.alignment = data_alignment_right
                    
                    # 根據課程狀態設定背景色
                    if idx == 0:  # 課程狀態欄位
                        if cell.value == "開課中":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif cell.value == "即將開課":
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        elif cell.value == "已結束":
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # 儲存 Excel 檔案
            workbook.save(file_path)
            
            QMessageBox.information(None, "成功", f"課程資料已成功匯出到：\n{file_path}")
            return True
            
        except PermissionError:
            QMessageBox.critical(None, "錯誤", "無法存取檔案，可能是檔案已開啟或沒有寫入權限")
            return False
        except Exception as e:
            QMessageBox.critical(None, "錯誤", f"匯出過程發生錯誤：\n{str(e)}")
            return False